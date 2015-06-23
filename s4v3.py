from s4v2 import *
import openpyxl
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.cell import get_column_letter

def save_spreadsheet(filename, data_sample):
	wb = Workbook() # shortcut for typing Workbook function
	ws = wb.active # shortcut for typing active workbook function and also, for finding the sheet in the workbook that we're working on, the active one.
	row_index = 1 # set the row index to 1, the starting point for excel, i.e. the upper left-hand corner
	for rows in data_sample: # iterate through the rows in the spreadsheet
		col_index = 1 # set the col index to 1 (starting point for excel, i.e. the upper left-hand corner)
		for field in rows:
			col_letter = get_column_letter(col_index) # use the imported get column letter function to get the letter of the column that we're working in.
			ws.cell('{}{}'.format(col_letter, row_index)).value = field # I'm not entirely sure what we're doing here because I haven't worked with these function before, but my guess is that we're writing the values in the field of the data sample into the current cell of the new workbook
			col_index += 1 # increase column index
		row_index += 1 # increase row index
	wb.save(filename)

kiton_ties = filter_col_by_string(data_from_csv, "brandName", "Kiton")
# save_spreadsheet("_data/s4-kiton.xlsx", kiton_ties)